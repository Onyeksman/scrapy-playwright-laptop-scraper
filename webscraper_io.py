import scrapy
from scrapy.crawler import CrawlerProcess
from scrapy_playwright.page import PageMethod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class LaptopsSpider(scrapy.Spider):
    name = "laptops_spider"
    start_urls = ["https://webscraper.io/test-sites/e-commerce/allinone/computers/laptops"]

    custom_settings = {
        "LOG_LEVEL": "INFO",
        "DOWNLOAD_HANDLERS": {
            "http": "scrapy_playwright.handler.ScrapyPlaywrightDownloadHandler",
            "https": "scrapy_playwright.handler.ScrapyPlaywrightDownloadHandler"
        },
        "TWISTED_REACTOR": "twisted.internet.asyncioreactor.AsyncioSelectorReactor",
        "PLAYWRIGHT_LAUNCH_OPTIONS": {"headless": True},
        "CONCURRENT_REQUESTS": 8,
        "AUTOTHROTTLE_ENABLED": True,
    }

    def __init__(self):
        self.products = []
        self.xlsx_file = "Product_Details.xlsx"

    def clean_text(self, text):
        if not text:
            return ""
        return text.replace('"', '').strip()

    def start_requests(self):
        for url in self.start_urls:
            yield scrapy.Request(
                url,
                meta={"playwright": True, "playwright_page_methods": [PageMethod("wait_for_selector", "div.product-wrapper.card-body")]}
            )

    def parse(self, response):
        product_blocks = response.css("div.product-wrapper.card-body")
        self.logger.info(f"Found {len(product_blocks)} product blocks on {response.url}")

        for product in product_blocks:
            title = product.css("a.title::attr(title)").get()
            price = product.css("span[itemprop='price']::text").get()
            description = product.css("p.description::text").get()
            reviews = product.css("span[itemprop='reviewCount']::text").get()

            self.products.append({
                "Title": self.clean_text(title) or "N/A",
                "Price": self.clean_text(price) or "N/A",
                "Description": self.clean_text(description) or "N/A",
                "Reviews": self.clean_text(reviews) or "N/A"
            })

        # Handle pagination if any
        next_page = response.css("ul.pagination li a[rel='next']::attr(href)").get()
        if next_page:
            yield response.follow(
                next_page,
                meta={"playwright": True, "playwright_page_methods": [PageMethod("wait_for_selector", "div.product-wrapper.card-body")]}
            )

    def closed(self, reason):
        # Remove duplicates
        seen = set()
        cleaned_products = []
        for item in self.products:
            key = (item["Title"], item["Price"])
            if key not in seen:
                seen.add(key)
                cleaned_products.append(item)

        # Excel output
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["Title", "Price", "Description", "Reviews"]
        ws.append(headers)

        for item in cleaned_products:
            ws.append([item["Title"], item["Price"], item["Description"], item["Reviews"]])

        # Excel formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        grey_font = Font(color="808080", italic=True)
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        border_side = Side(border_style="medium", color="000000")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if r_idx % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill
            for idx, cell in enumerate(row):
                if str(cell.value).strip() == "N/A":
                    cell.font = grey_font
                cell.border = border
                # Numeric formatting
                if idx == 1:  # Price column
                    try:
                        val = float(str(cell.value).replace("$", "").replace(",", ""))
                        cell.value = val
                        cell.number_format = "$#,##0.00"
                    except:
                        pass
                if idx == 3:  # Reviews column
                    try:
                        val = int(str(cell.value).replace("$", "").replace(",", ""))
                        cell.value = val
                        cell.number_format = "#,##0"
                    except:
                        pass

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 80)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Metadata
        last_row = ws.max_row + 2
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=last_row, column=1, value="📊 Sourced from (https://webscraper.io/test-sites/e-commerce/allinone/computers/laptops)").font = grey_font
        ws.cell(row=last_row+1, column=1, value=f"Generated on: {timestamp}").font = grey_font

        wb.save(self.xlsx_file)
        self.logger.info(f"Wrote {len(cleaned_products)} products to {self.xlsx_file} ({reason})")


# Run spider
if __name__ == "__main__":
    process = CrawlerProcess()
    process.crawl(LaptopsSpider)
    process.start()
